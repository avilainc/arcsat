from fastapi import APIRouter, HTTPException, Response
from typing import List, Optional
from datetime import datetime, timedelta
from bson import ObjectId
from ..database import customers_collection, deals_collection, interactions_collection, activities_collection
import csv
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter()

@router.get("/reports/customers/csv")
async def export_customers_csv(status: Optional[str] = None):
    """Exportar clientes para CSV"""
    try:
        query = {}
        if status:
            query["status"] = status

        customers = await customers_collection.find(query).to_list(10000)

        # Criar CSV em memória
        output = io.StringIO()
        writer = csv.writer(output)

        # Cabeçalhos
        headers = [
            'ID', 'Nome', 'Email', 'Telefone', 'Empresa', 'CNPJ', 'Status',
            'Cidade', 'UF', 'Categoria', 'Origem', 'Responsável', 'Score',
            'Valor Contrato', 'Data Cadastro'
        ]
        writer.writerow(headers)

        # Dados
        for customer in customers:
            writer.writerow([
                str(customer['_id']),
                customer.get('name', ''),
                customer.get('email', ''),
                customer.get('phone', ''),
                customer.get('company', ''),
                customer.get('cnpj', ''),
                customer.get('status', ''),
                customer.get('municipio', ''),
                customer.get('uf', ''),
                customer.get('categoria', ''),
                customer.get('origem', ''),
                customer.get('responsavel', ''),
                customer.get('score', 0),
                customer.get('valor_contrato', 0),
                customer.get('created_at', datetime.now()).strftime('%d/%m/%Y')
            ])

        csv_content = output.getvalue()
        output.close()

        return Response(
            content=csv_content.encode('utf-8-sig'),
            media_type='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename=clientes_{datetime.now().strftime("%Y%m%d")}.csv'
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao exportar CSV: {str(e)}")


@router.get("/reports/customers/excel")
async def export_customers_excel(status: Optional[str] = None):
    """Exportar clientes para Excel com formatação"""
    try:
        query = {}
        if status:
            query["status"] = status

        customers = await customers_collection.find(query).to_list(10000)

        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Cabeçalhos
        headers = [
            'ID', 'Nome', 'Email', 'Telefone', 'Empresa', 'CNPJ', 'Status',
            'Cidade', 'UF', 'Categoria', 'Origem', 'Responsável', 'Score',
            'Valor Contrato', 'Data Cadastro'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Dados
        for row, customer in enumerate(customers, start=2):
            ws.cell(row=row, column=1, value=str(customer['_id']))
            ws.cell(row=row, column=2, value=customer.get('name', ''))
            ws.cell(row=row, column=3, value=customer.get('email', ''))
            ws.cell(row=row, column=4, value=customer.get('phone', ''))
            ws.cell(row=row, column=5, value=customer.get('company', ''))
            ws.cell(row=row, column=6, value=customer.get('cnpj', ''))
            ws.cell(row=row, column=7, value=customer.get('status', ''))
            ws.cell(row=row, column=8, value=customer.get('municipio', ''))
            ws.cell(row=row, column=9, value=customer.get('uf', ''))
            ws.cell(row=row, column=10, value=customer.get('categoria', ''))
            ws.cell(row=row, column=11, value=customer.get('origem', ''))
            ws.cell(row=row, column=12, value=customer.get('responsavel', ''))
            ws.cell(row=row, column=13, value=customer.get('score', 0))
            ws.cell(row=row, column=14, value=customer.get('valor_contrato', 0))
            ws.cell(row=row, column=15, value=customer.get('created_at', datetime.now()).strftime('%d/%m/%Y'))

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Salvar em memória
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return Response(
            content=excel_file.getvalue(),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=clientes_{datetime.now().strftime("%Y%m%d")}.xlsx'
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao exportar Excel: {str(e)}")


@router.get("/reports/pipeline/pdf")
async def export_pipeline_pdf():
    """Exportar relatório do pipeline em PDF"""
    try:
        # Buscar dados
        pipeline_stages = ["Prospecção", "Qualificação", "Proposta", "Negociação", "Fechamento"]
        pipeline_data = []

        for stage in pipeline_stages:
            deals = await deals_collection.find({"stage": stage, "status": "open"}).to_list(1000)
            total_value = sum(d.get('value', 0) for d in deals)
            pipeline_data.append([
                stage,
                len(deals),
                f"R$ {total_value:,.2f}"
            ])

        # Criar PDF
        pdf_file = io.BytesIO()
        doc = SimpleDocTemplate(pdf_file, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=1  # Center
        )
        title = Paragraph("Relatório Pipeline de Vendas", title_style)
        elements.append(title)

        # Data
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=1
        )
        date_text = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", date_style)
        elements.append(date_text)
        elements.append(Spacer(1, 20))

        # Tabela
        table_data = [['Estágio', 'Quantidade', 'Valor Total']] + pipeline_data
        table = Table(table_data, colWidths=[2*inch, 1.5*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)

        # Stats adicionais
        elements.append(Spacer(1, 30))
        total_deals = sum(row[1] for row in pipeline_data)
        total_value = sum(float(row[2].replace('R$ ', '').replace(',', '')) for row in pipeline_data)

        stats_text = f"""
        <b>Resumo Geral:</b><br/>
        Total de Deals Abertos: {total_deals}<br/>
        Valor Total em Negociação: R$ {total_value:,.2f}
        """
        stats_para = Paragraph(stats_text, styles['Normal'])
        elements.append(stats_para)

        doc.build(elements)
        pdf_file.seek(0)

        return Response(
            content=pdf_file.getvalue(),
            media_type='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename=pipeline_{datetime.now().strftime("%Y%m%d")}.pdf'
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar PDF: {str(e)}")


@router.get("/reports/activities")
async def export_activities_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None
):
    """Relatório de atividades com filtros"""
    try:
        query = {}

        if start_date:
            query["created_at"] = {"$gte": datetime.fromisoformat(start_date)}
        if end_date:
            if "created_at" in query:
                query["created_at"]["$lte"] = datetime.fromisoformat(end_date)
            else:
                query["created_at"] = {"$lte": datetime.fromisoformat(end_date)}
        if status:
            query["status"] = status

        activities = await activities_collection.find(query).to_list(10000)

        # Estatísticas
        total = len(activities)
        by_type = {}
        by_status = {}

        for activity in activities:
            act_type = activity.get('activity_type', 'outros')
            act_status = activity.get('status', 'unknown')

            by_type[act_type] = by_type.get(act_type, 0) + 1
            by_status[act_status] = by_status.get(act_status, 0) + 1

        return {
            "period": {
                "start": start_date or "início",
                "end": end_date or "hoje"
            },
            "total_activities": total,
            "by_type": by_type,
            "by_status": by_status,
            "activities": [
                {
                    "id": str(act["_id"]),
                    "title": act.get("title"),
                    "type": act.get("activity_type"),
                    "status": act.get("status"),
                    "customer_id": act.get("customer_id"),
                    "created_at": act.get("created_at"),
                    "due_date": act.get("due_date")
                }
                for act in activities[:100]  # Limitar a 100 na resposta
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório: {str(e)}")


@router.get("/reports/interactions")
async def export_interactions_report(customer_id: Optional[str] = None, days: int = 30):
    """Relatório de interações dos últimos N dias"""
    try:
        start_date = datetime.now() - timedelta(days=days)
        query = {"created_at": {"$gte": start_date}}

        if customer_id:
            query["customer_id"] = customer_id

        interactions = await interactions_collection.find(query).sort("created_at", -1).to_list(10000)

        # Estatísticas
        by_type = {}
        by_result = {}

        for interaction in interactions:
            int_type = interaction.get('tipo', 'outros')
            int_result = interaction.get('resultado', 'pendente')

            by_type[int_type] = by_type.get(int_type, 0) + 1
            by_result[int_result] = by_result.get(int_result, 0) + 1

        return {
            "period_days": days,
            "total_interactions": len(interactions),
            "by_type": by_type,
            "by_result": by_result,
            "interactions": [
                {
                    "id": str(i["_id"]),
                    "customer_id": i.get("customer_id"),
                    "tipo": i.get("tipo"),
                    "titulo": i.get("titulo"),
                    "resultado": i.get("resultado"),
                    "data": i.get("data"),
                    "created_at": i.get("created_at")
                }
                for i in interactions[:100]
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório: {str(e)}")


@router.get("/reports/conversion-funnel")
async def conversion_funnel_report():
    """Relatório detalhado do funil de conversão"""
    try:
        # Contar por status
        total = await customers_collection.count_documents({})
        leads = await customers_collection.count_documents({"status": "lead"})
        prospects = await customers_collection.count_documents({"status": "prospect"})
        clientes = await customers_collection.count_documents({"status": "cliente"})

        # Calcular conversões
        lead_to_prospect = (prospects / leads * 100) if leads > 0 else 0
        prospect_to_client = (clientes / prospects * 100) if prospects > 0 else 0
        lead_to_client = (clientes / leads * 100) if leads > 0 else 0

        # Tempo médio de conversão
        pipeline = [
            {"$match": {"status": "cliente"}},
            {"$addFields": {
                "conversion_time": {
                    "$divide": [
                        {"$subtract": ["$updated_at", "$created_at"]},
                        86400000  # Convert to days
                    ]
                }
            }},
            {"$group": {
                "_id": None,
                "avg_days": {"$avg": "$conversion_time"}
            }}
        ]
        conversion_time = await customers_collection.aggregate(pipeline).to_list(1)
        avg_conversion_days = conversion_time[0]["avg_days"] if conversion_time else 0

        return {
            "funnel_stages": [
                {
                    "stage": "Leads",
                    "count": leads,
                    "percentage": 100,
                    "conversion_to_next": lead_to_prospect
                },
                {
                    "stage": "Prospects",
                    "count": prospects,
                    "percentage": (prospects/leads*100) if leads > 0 else 0,
                    "conversion_to_next": prospect_to_client
                },
                {
                    "stage": "Clientes",
                    "count": clientes,
                    "percentage": (clientes/leads*100) if leads > 0 else 0,
                    "conversion_to_next": 0
                }
            ],
            "overall_conversion": round(lead_to_client, 2),
            "avg_conversion_time_days": round(avg_conversion_days, 1),
            "total_customers": total
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório: {str(e)}")
